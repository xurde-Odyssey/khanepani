"""
One-off migration script: import the historical Excel workbook into
`daily_entries` for cutover. Not part of the deployed app — run locally once.

Usage:
    pip install openpyxl supabase python-dotenv
    python scripts/migrate_excel.py path/to/legacy_workbook.xlsx <project_id>

This is a STARTING POINT, not a finished importer — the exact row/column
layout (which row each variable block starts on, which columns are days vs.
Total) needs to be adjusted to match the real file. Ask the client for the
workbook first, open it, and fill in SHEET_LAYOUT below before running.
"""
import sys
import os
from datetime import date
import openpyxl
from supabase import create_client

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_ROLE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]  # service role, NOT anon key — bypasses RLS for the import

BS_MONTHS = [
    "Shrawan", "Bhadra", "Ashoj", "Karthik", "Mangsir", "Poush",
    "Magh", "Falgun", "Chaitra", "Baishak", "Jestha", "Ashar",
]

# TODO: fill this in against the real workbook. Example shape:
# each BS month is one worksheet tab; each variable is a contiguous block of
# rows (one row per pump), day columns run left-to-right, Total is the last column.
SHEET_LAYOUT = {
    # "Shrawan": {
    #     "operating_hours": {"start_row": 3, "n_pumps": 6},
    #     "flowmeter_start_unit": {"start_row": 10, "n_pumps": 6},
    #     ...
    # },
}


def bs_to_gregorian(bs_year: int, bs_month: str, bs_day: int) -> date:
    """Plug in whatever BS conversion utility/table you use for the import
    (e.g. a Python port of the same library used in src/lib/bsCalendar.ts,
    or a lookup table). Left unimplemented here since it's data-source
    specific — raise until wired up so bad dates don't silently import."""
    raise NotImplementedError("Wire up a BS->AD converter before running this script.")


def main():
    if len(sys.argv) != 3:
        print("Usage: python migrate_excel.py <workbook.xlsx> <project_id>")
        sys.exit(1)

    workbook_path, project_id = sys.argv[1], sys.argv[2]
    supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    rows_to_insert = {}  # (pump_no, entry_date) -> dict of fields

    for bs_month, layout in SHEET_LAYOUT.items():
        if bs_month not in wb.sheetnames:
            print(f"WARNING: sheet '{bs_month}' not found, skipping")
            continue
        ws = wb[bs_month]
        for variable, meta in layout.items():
            start_row = meta["start_row"]
            n_pumps = meta["n_pumps"]
            for i in range(n_pumps):
                pump_row = start_row + i
                pump_no = i + 1  # adjust if pump numbering isn't sequential from row 1
                for day in range(1, 32):
                    col = 1 + day  # adjust to match actual day-column offset
                    cell = ws.cell(row=pump_row, column=col)
                    if cell.value is None:
                        continue
                    # bs_year needs to come from somewhere — a fixed constant for this
                    # cutover, or a cell in the sheet. Adjust as needed.
                    bs_year = 2081
                    key = (pump_no, bs_year, bs_month, day)
                    rows_to_insert.setdefault(key, {})[variable] = float(cell.value)

    # Look up pump UUIDs by pump_no within the project.
    pumps = supabase.table("pumps").select("id, pump_no").eq("project_id", project_id).execute().data
    pump_id_by_no = {p["pump_no"]: p["id"] for p in pumps}

    payload = []
    for (pump_no, bs_year, bs_month, bs_day), fields in rows_to_insert.items():
        if pump_no not in pump_id_by_no:
            print(f"WARNING: no pump found for pump_no={pump_no}, skipping")
            continue
        entry_date = bs_to_gregorian(bs_year, bs_month, bs_day)
        payload.append({
            "pump_id": pump_id_by_no[pump_no],
            "entry_date": entry_date.isoformat(),
            "bs_year": bs_year,
            "bs_month": bs_month,
            "bs_day": bs_day,
            **fields,
        })

    print(f"Prepared {len(payload)} rows. Inserting in batches of 500…")
    for i in range(0, len(payload), 500):
        batch = payload[i:i + 500]
        supabase.table("daily_entries").upsert(batch, on_conflict="pump_id,entry_date").execute()
    print("Done.")


if __name__ == "__main__":
    main()
